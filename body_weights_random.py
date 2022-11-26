from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import random

wb = Workbook()

ws = wb.active
ws.title = "Body Weights"

ws['A1'] = "Counter"
ws['B1'] = "Body Weight"

i = 1

while i <= 100:
    ws.cell(row= 1 + i, column= 1).value = i
    ws.cell(row= 1 + i, column= 2).value = round(random.uniform(120, 250), 2)
    i = i + 1

wb.save('body_weights.xlsx')
