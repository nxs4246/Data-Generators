from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import random

def format(num):
    if num < 10:
        num = "0" + str(num)
    return num

wb = Workbook()

ws = wb.active
ws.title = "Time Interval"

ws['A1'] = "Time"
ws['B1'] = "Time in Seconds"

start_time = 32400

i = 1
while i <= 100:
    start_time = start_time + random.randint(200, 350)
    ws.cell(row= 1 + i, column= 2).value = start_time

    hour = int(start_time / 3600)
    minute = int((start_time % 3600) / 60)
    second = int((start_time % 3600) % 60)

    hour = format(hour)
    minute = format(minute)
    second = format(second)

    ws.cell(row= 1 + i, column= 1).value = ("%s%s%s%s%s" % (hour, ":", minute, ":", second))
    i = i + 1

wb.save('time_interval.xlsx')
